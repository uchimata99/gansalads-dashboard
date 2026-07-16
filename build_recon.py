#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
בניית הצלבת "קניות מול ייצור" — לכל חומר גלם: כמה נצרך תאורטית לפי הייצור בפועל
(× עץ המוצר) מול כמה נקנה בפועל. קוד מחויב, לא צ'אט.

מקורות
------
- עצי מוצר (`--trees "עצי מוצר.xlsx"`): אינדקס עצי מוצר של מחשבשבת. כל עץ =
  מוצר מוגמר; הרמות (עמודות 0..5) מגדירות מרכיבים והכמות פר יחידה. פורסים
  רקורסיבית לעלים (חומרי גלם) עם מקדם = מכפלת הכמויות בשרשרת.
- ייצור בפועל: לשונית DATA (מפתח `PROD_DETAIL`) בגיליון הייצור — אריזות פר
  מוצר/שבוע. נקרא דרך חשבון השירות.
- קניות: `mat_prices.json` (פלט build_veg_prices.py) — ק"ג שנקנו פר חומר לשנה.

זיהוי עצי-אצווה
--------------
חלק מהעצים במחשבשבת מוזנים בכמויות **אצווה** ולא פר-יחידה (למשל איולי לימון:
17,604 ק"ג עמילן "ליחידה"). מוצר כזה מזוהה כשמקדם כלשהו > 6× משקל האריזה,
ו**מוחרג כולו** מהצריכה התאורטית; נספר `nBatch` פר חומר כדי לסמן כיסוי חלקי.
עצי-אצווה לא נמחקים — הם ממתינים לסמנטיקת האצווה ממיכל (ואז מנרמלים).

פלט: recon.json (לא מחויב — נקלט ללשונית RECON בלבד).
"""
import argparse
import json
import os
import re
from datetime import date


def pkg_kg(name):
    """משקל אריזה מתוך שם המוצר: 'X ק\"ג' / 'X גרם' -> ק\"ג (ברירת מחדל 3)."""
    m = re.search(r'([\d.]+)\s*ק"?ג', name)
    if m:
        return float(m.group(1))
    m = re.search(r'([\d.]+)\s*גרם', name)
    if m:
        return float(m.group(1)) / 1000.0
    return 3.0


def read_trees(path):
    """עצי מוצר -> {key:{'name','pkg','comp':[(level,key,name,qty)]}}."""
    from openpyxl import load_workbook
    ws = load_workbook(path, read_only=True, data_only=True).active
    trees, cur = {}, None
    for row in ws.iter_rows(values_only=True):
        key = str(row[1] or '').strip()
        name = str(row[2] or '').strip()
        typ = str(row[3] or '').strip()
        if not key:
            continue
        lvl = qty = None
        for L in range(6):                     # עמודות 4..9 = רמות 0..5
            v = row[4 + L]
            if v is not None and str(v).strip() != '':
                try:
                    qty = float(v)
                except (TypeError, ValueError):
                    break
                lvl = L
                break
        if lvl is None:
            continue
        if typ == 'עץ יצור' and lvl == 0:
            cur = {'key': key, 'name': name, 'pkg': pkg_kg(name), 'comp': []}
            trees[key] = cur
            continue
        if cur is not None:
            cur['comp'].append((lvl, key, name, qty))
    return trees


def flatten(tree):
    """פירוק עץ לעלים ישירים (מרכיב שהוא בעצמו עץ נשאר לפריסה צולבת אחר-כך)."""
    out, path = {}, []
    comp = tree['comp']
    for i, (lvl, key, name, qty) in enumerate(comp):
        while path and path[-1][0] >= lvl:
            path.pop()
        mult = (path[-1][1] if path else 1.0) * qty
        nxt = comp[i + 1][0] if i + 1 < len(comp) else None
        if nxt is not None and nxt > lvl:      # יש תת-מרכיבים -> צומת
            path.append((lvl, mult))
        else:                                  # עלה
            out[key] = out.get(key, 0.0) + mult
    return out


def cross_expand(flat, key, seen=None):
    """מרכיב שהוא עץ בעצמו -> מפורק לחומרי הגלם שלו."""
    seen = seen or set()
    if key in seen:
        return {}
    seen = seen | {key}
    out = {}
    for k, q in flat.get(key, {}).items():
        if k in flat and k != key:
            for kk, qq in cross_expand(flat, k, seen).items():
                out[kk] = out.get(kk, 0.0) + q * qq
        else:
            out[k] = out.get(k, 0.0) + q
    return out


def read_production(key_json, sheet_id, max_week):
    """PROD_DETAIL מלשונית DATA -> {product_name: units} עד max_week (כולל)."""
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    creds = Credentials.from_service_account_file(
        key_json, scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"])
    sh = build("sheets", "v4", credentials=creds, cache_discovery=False).spreadsheets()
    rows = sh.values().get(spreadsheetId=sheet_id, range="DATA!A:A").execute().get("values", [])
    D = json.loads("".join(r[0] for r in rows if r))
    units = {}
    for name, wk in D.get("PROD_DETAIL", {}).items():
        for w, days in wk.items():
            if int(w) <= max_week:
                for _, t in days.items():
                    units[name] = units.get(name, 0.0) + t[0]
    return units


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--trees", required=True, help='"עצי מוצר.xlsx"')
    ap.add_argument("--mat", required=True, help="mat_prices.json (פלט build_veg_prices.py)")
    ap.add_argument("--key", default=os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"),
                    help="מפתח חשבון שירות (לקריאת הייצור מהגיליון)")
    ap.add_argument("--prod-sheet", default="1_rJ8lLYNme8RM83ws1pFOGij0HPAVAUFefBDYl2_B5A",
                    help="גיליון הייצור (לשונית DATA)")
    ap.add_argument("--weeks", type=int, default=26, help="עד שבוע ייצור זה (≈ התקופה)")
    ap.add_argument("--year", default="2026", help="שנת הקניות להשוואה (מ-mat_prices)")
    ap.add_argument("--out", default="recon.json")
    ap.add_argument("--date", default=None)
    args = ap.parse_args()
    if not args.key:
        raise SystemExit("חסר מפתח חשבון שירות: --key או GOOGLE_APPLICATION_CREDENTIALS")

    trees = read_trees(args.trees)
    flat = {k: flatten(t) for k, t in trees.items()}
    bom = {k: cross_expand(flat, k) for k in flat}              # מוצר -> {חומר: מקדם}

    # עצי-אצווה: מקדם כלשהו > 6× משקל אריזה -> העץ כולו מוחרג
    batch = {k for k in bom if any(c > max(6 * trees[k]['pkg'], 3) for c in bom[k].values())}

    mat = json.loads(open(args.mat, encoding="utf-8").read())
    items = {it['key']: it for it in mat['items']}
    name2key = {it['name']: it['key'] for it in mat['items']}   # שם מוצר-גלם -> מפתח
    # שמות מוצרים מוגמרים (מהעצים) לצורך מיפוי הייצור
    fp_name2key = {t['name']: k for k, t in trees.items()}

    units = read_production(args.key, args.prod_sheet, args.weeks)
    unmapped = sorted([n for n in units if n not in fp_name2key and units[n] > 30])

    # צריכה תאורטית פר חומר גלם (רק ממוצרים לא-אצווה שמופו)
    theo, contrib, nbatch = {}, {}, {}
    for pname, u in units.items():
        pk = fp_name2key.get(pname)
        if not pk:
            continue
        for mk, coef in bom.get(pk, {}).items():
            if pk in batch:
                nbatch[mk] = nbatch.get(mk, 0) + 1
                continue
            theo[mk] = theo.get(mk, 0.0) + u * coef
            cd = contrib.setdefault(mk, {})
            cd[pname] = cd.get(pname, 0.0) + u * coef

    def purch_year(k):
        it = items.get(k)
        if not it:
            return None
        a = [x for x in it.get('annual', []) if x['year'] == args.year]
        return a[0]['kg'] if a else 0.0

    out_items = []
    allkeys = set(theo) | {it['key'] for it in mat['items'] if purch_year(it['key'])}
    for k in allkeys:
        if not k.startswith('70'):
            continue
        t = round(theo.get(k, 0.0), 1)
        p = purch_year(k)
        if (t or 0) < 20 and (p or 0) < 20:
            continue
        ratio = (p / t) if (t and p is not None) else None
        top = sorted(contrib.get(k, {}).items(), key=lambda x: -x[1])[:4]
        out_items.append({
            'key': k, 'name': items.get(k, {}).get('name', k),
            'grp': items.get(k, {}).get('grp'),
            'theo': t, 'purch': None if p is None else round(p, 1),
            'ratio': None if ratio is None else round(ratio, 3),
            'nBatch': nbatch.get(k, 0),
            'top': [{'name': n, 'kg': round(v, 1)} for n, v in top]})
    out_items.sort(key=lambda x: -(x['theo'] or 0))

    payload = {
        'generatedAt': args.date or date.today().isoformat(),
        'year': args.year,
        'weeks': args.weeks,
        'groups': mat.get('groups', {}),
        'items': out_items,
        'batchProducts': sorted(trees[k]['name'] for k in batch),
        'unmappedProduction': unmapped,
        'totals': {'nItems': len(out_items), 'nBatch': len(batch),
                   'nUnmapped': len(unmapped)},
    }
    json.dump(payload, open(args.out, "w", encoding="utf-8"),
              ensure_ascii=False, separators=(",", ":"))
    print(f"נכתב {args.out} | חומרים: {len(out_items)} | עצי-אצווה שהוחרגו: {len(batch)} | "
          f"מוצרי ייצור לא ממופים: {len(unmapped)}")
    print("\nדוגמאות (חומר | תאורטי | נקנה | יחס):")
    for it in out_items[:14]:
        r = '—' if it['ratio'] is None else f"{it['ratio']*100:.0f}%"
        print(f"  {it['name'][:26]:<28} {it['theo']:>10,.0f} | {str(it['purch']):>10} | {r:>6}"
              + (f"  ⚠️{it['nBatch']} אצווה" if it['nBatch'] else ''))


if __name__ == "__main__":
    main()
