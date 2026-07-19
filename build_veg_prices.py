#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
בניית ניתוח מחירי חומרי גלם וירקות מכרטסות המלאי של מחשבשבת (קוד מחויב, לא צ'אט).

מקור הנתונים
------------
תיקיית דרייב "קניית ירקות":
- "כרטסת מלאי חומרי גלם 25_26.xlsx" — כרטסת רב-פריטים אחת לכל חומרי הגלם
  (פורמט מחשבשבת: כרטסת מלאי לפי תאריכי רישום; בלוק לכל פריט: שורת כותרת
  [שם, מפתח, קוד מיון], תנועות, ושורת "סה\"כ מפתח פריט").
- נתמכות גם כרטסות של פריט בודד (הפורמט הישן) וייצוא טקסט/CSV.

סוגי מסמך: לניתוח הרכש נספרות שורות "חשבונית רכש" בלבד (מחיר נטו ₪ לק"ג,
כניסה בק"ג). "חשבונית מס"/"ריכוז"/"קבלה" = מכירת חומר גלם החוצה — לא רכש.
"החזרה" = החזרת רכש (נאספת לקטע returns). בדיקת שלמות: סכום כל הכניסות/יציאות
(מכל סוגי המסמכים) מול שורת הסה"כ של הפריט.

מה זה מייצר
-----------
--out veg_prices.json — ניתוח הירקות (טאב "מחירי ירקות"): הירקות הטריים
  שברשימת VEG_FRESH + מוצרי משפחת העגבניות (TOMATO_FAMILY, מסומנים fam).
--mat-out mat_prices.json — ניתוח כל חומרי הגלם (טאב "חומרי גלם"): כל פריט
  עם רכש, כולל קוד קבוצה (קוד מיון). תוויות הקבוצות והספקים מועשרות בקליטה
  מקטלוג הרכש (לא נשמרות בקוד).

מבנה פריט (זהה בשניהם): {key,name,grp,totalKg,totalSpend,avgPrice,minPrice,
  maxPrice,nTxn,firstMonth,lastMonth,monthly:[{m,kg,spend,avg,min,max,n}],
  annual:[{year,kg,spend,avg,min,max,n,bySupplier}],bySupplier,fam?}
מחיר = ממוצע משוקלל בק"ג: Σ(מחיר×ק"ג)÷Σק"ג.

אבטחה: מחירי הקנייה רגישים. veg_prices.json / mat_prices.json **אינם מחויבים
למאגר** — נבנים ונקלטים לגיליון בלבד (שניהם ב-.gitignore).

הרצה
----
  python3 build_veg_prices.py "כרטסת מלאי חומרי גלם *.xlsx" \
      --out veg_prices.json --mat-out mat_prices.json
"""
import argparse
import glob
import json
import re
import sys
from datetime import date, datetime

DOC_PURCHASE = "חשבונית רכש"
RETURN_HINT = "החזר"   # "החזרה" / "החזרת רכש"

# הירקות הטריים של טאב "מחירי ירקות" (מפתח פריט, קבוצה 206)
VEG_FRESH = ["70527", "70526", "70523", "70520",      # עגבניות, חציל, פלפל חריף, פלפל אדום
             "70522", "70521", "70528", "70538"]      # פלפל ירוק, פלפל צהוב, מלפפון, קישואים
# משפחת העגבניות — מוצרי צ'אם שמצטרפים להשוואת הכמויות של עגבניות (בהפרדה)
TOMATO_FAMILY = {"70009": "70527",    # קוביות עגבניות -> בסיס עגבניות
                 "70611": "70527"}    # עגבניות מרוסקות -> בסיס עגבניות

SUPPLIER_LABELS = {}   # שמות ספקים/קבוצות מועשרים בקליטה מקטלוג הרכש — לא בקוד

# נרמול יחידות מידה — פריטים שהיחידה שלהם השתנתה במחשבשבת באמצע התקופה.
# מפתח פריט -> רשימת כללים: תנועה לפני before (YYYY-MM-DD) מוכפלת: מחיר ÷ factor,
# כמות × factor. כל כלל מגיע מהסבר מפורש של מיכל — לא מנחשים.
UNIT_FIXES = {
    "70401": [{"before": "2025-08-01", "factor": 5.0,
               "why": "שמן קנולה נקנה בבקבוקי 5 ליטר עד אמצע 2025 (₪29–30 לבקבוק); "
                      "מאז נקנה בתפזורת לליטר. מיכל, 15/07/26 — המחיר לליטר כמעט זהה."}],
}


def _apply_unit_fixes(it):
    """מנרמל תנועות לפי UNIT_FIXES; מחזיר את מספר התנועות שתוקנו."""
    rules = UNIT_FIXES.get(str(it['key']))
    if not rules:
        return 0
    fixed = 0
    for t in it['txns']:
        m = _month(t['date'])
        day = None
        if isinstance(t['date'], (datetime, date)):
            day = t['date'].strftime('%Y-%m-%d')
        elif m:
            mm = re.match(r'\s*(\d{1,2})/(\d{1,2})/(\d{2,4})', str(t['date']))
            if mm:
                yy = int(mm.group(3)); yy = yy + 2000 if yy < 100 else yy
                day = f"{yy:04d}-{int(mm.group(2)):02d}-{int(mm.group(1)):02d}"
        if day is None:
            continue
        for r in rules:
            if day < r['before']:
                t['price'] = t['price'] / r['factor']
                t['inn'] = t['inn'] * r['factor']
                t['out'] = t['out'] * r['factor']
                fixed += 1
    return fixed


def _num(x):
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace('"', '').replace(',', '')
    if s in ('', '—'):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _month(d):
    """תאריך (datetime או DD/MM/YY) -> 'YYYY-MM'."""
    if isinstance(d, (datetime, date)):
        return f"{d.year:04d}-{d.month:02d}"
    m = re.match(r'\s*(\d{1,2})/(\d{1,2})/(\d{2,4})', str(d or ''))
    if not m:
        return None
    yy = int(m.group(3))
    return f"{(yy + 2000 if yy < 100 else yy):04d}-{int(m.group(2)):02d}"


def _datestr(d):
    """תאריך -> 'DD/MM/YY' לתצוגה."""
    if isinstance(d, (datetime, date)):
        return f"{d.day:02d}/{d.month:02d}/{d.year % 100:02d}"
    return str(d or '')


# ------------------------- קריאת כרטסת -------------------------
def read_ledger_xlsx(path):
    """כרטסת מחשבשבת (פריט בודד או רב-פריטים) -> רשימת פריטים.
    פריט: {key,name,grp,txns:[{doc,code,date,price,inn,out}],total_in,total_out}"""
    from openpyxl import load_workbook
    ws = load_workbook(path, read_only=True, data_only=True).active
    items, cur = [], None
    ci = None   # מפת עמודות לפי שורת הכותרת
    for row in ws.iter_rows(values_only=True):
        vals = [c for c in row]
        s0 = str(vals[0] or '').strip()
        s1 = str(vals[1] or '').strip() if len(vals) > 1 else ''
        s2 = str(vals[2] or '').strip() if len(vals) > 2 else ''
        if ci is None:
            svals = [str(c).strip() if c is not None else '' for c in vals]
            if 'מחיר נטו' in svals and 'כניסה' in svals:
                ci = {k: svals.index(k) for k in
                      ['סוג מסמך', 'לקוח/ספק', 'תאריך', 'מחיר נטו', 'כניסה', 'יציאה'] if k in svals}
            continue
        if s0.startswith('סה'):                                  # שורות סיכום
            if 'מפתח פריט' in s0 and cur is not None:
                cur['total_in'] = _num(vals[ci['כניסה']])
                cur['total_out'] = _num(vals[ci['יציאה']]) if 'יציאה' in ci else 0.0
            continue
        if s0 and re.match(r'^\d{4,6}$', s1) and re.match(r'^\d{3}$', s2):   # כותרת פריט
            cur = {'key': s1, 'name': s0, 'grp': s2, 'txns': [],
                   'total_in': None, 'total_out': None}
            items.append(cur)
            continue
        doc = str(vals[ci['סוג מסמך']] or '').strip() if len(vals) > ci['סוג מסמך'] else ''
        if doc and cur is not None:
            cur['txns'].append({
                'doc': doc,
                'code': str(vals[ci['לקוח/ספק']] or '').strip(),
                'date': vals[ci['תאריך']],
                'price': _num(vals[ci['מחיר נטו']]),
                'inn': _num(vals[ci['כניסה']]),
                'out': _num(vals[ci['יציאה']]) if 'יציאה' in ci else 0.0})
    return items


def read_ledger_text(path):
    """ייצוא טקסט/CSV של כרטסת פריט בודד (הפורמט הישן) -> רשימת פריט אחד."""
    key = name = None
    total_kg = None
    txns = []
    for line in open(path, encoding='utf-8'):
        f = [c.strip() for c in line.rstrip('\n').split(',')]
        if not f:
            continue
        if f[0].startswith('סהכ מפתח פריט') or f[0].startswith('סה"כ מפתח פריט'):
            key = key or (f[1] if len(f) > 1 else None)
            name = name or (f[2] if len(f) > 2 else None)
            nums = [c for c in f if re.match(r'^-?\d+(\.\d+)?$', c or '')]
            if len(nums) >= 3:
                total_kg = _num(nums[-3])
            continue
        if len(f) >= 15 and f[0] == '' and f[1] == '' and f[2] == '' and f[4]:
            txns.append({'doc': f[4], 'code': f[5], 'date': f[10], 'price': _num(f[12]),
                         'inn': _num(f[14]), 'out': _num(f[15]) if len(f) > 15 else 0.0})
    return [{'key': str(key), 'name': name, 'grp': '206', 'txns': txns,
             'total_in': total_kg, 'total_out': 0.0}]


# ------------------------- אגרגציה -------------------------
def build_item(it):
    """פריט כרטסת -> מבנה ניתוח + שורות החזרה. רכש = 'חשבונית רכש' בלבד."""
    purch = [t for t in it['txns'] if t['doc'] == DOC_PURCHASE and t['inn'] > 0]
    returns = [t for t in it['txns'] if RETURN_HINT in t['doc']]

    by_month, by_sup, by_year, by_ys = {}, {}, {}, {}
    tot_kg = tot_spend = 0.0
    pmin = pmax = None
    for t in purch:
        m = _month(t['date'])
        yr = (m or '?').split('-')[0]
        kg, sp = t['inn'], t['price'] * t['inn']
        tot_kg += kg
        tot_spend += sp
        pmin = t['price'] if pmin is None else min(pmin, t['price'])
        pmax = t['price'] if pmax is None else max(pmax, t['price'])
        d = by_month.setdefault(m, {'kg': 0.0, 'spend': 0.0, 'min': t['price'], 'max': t['price'], 'n': 0})
        d['kg'] += kg; d['spend'] += sp; d['n'] += 1
        d['min'] = min(d['min'], t['price']); d['max'] = max(d['max'], t['price'])
        s = by_sup.setdefault(t['code'], {'kg': 0.0, 'spend': 0.0, 'n': 0})
        s['kg'] += kg; s['spend'] += sp; s['n'] += 1
        y = by_year.setdefault(yr, {'kg': 0.0, 'spend': 0.0, 'min': t['price'], 'max': t['price'], 'n': 0})
        y['kg'] += kg; y['spend'] += sp; y['n'] += 1
        y['min'] = min(y['min'], t['price']); y['max'] = max(y['max'], t['price'])
        ys = by_ys.setdefault((yr, t['code']), {'kg': 0.0, 'spend': 0.0, 'n': 0})
        ys['kg'] += kg; ys['spend'] += sp; ys['n'] += 1

    monthly = [{'m': m, 'kg': round(v['kg'], 1), 'spend': round(v['spend'], 1),
                'avg': round(v['spend'] / v['kg'], 3) if v['kg'] else None,
                'min': round(v['min'], 3), 'max': round(v['max'], 3), 'n': v['n']}
               for m, v in sorted(by_month.items())]
    bysup = [{'code': c, 'kg': round(v['kg'], 1), 'spend': round(v['spend'], 1),
              'avg': round(v['spend'] / v['kg'], 3) if v['kg'] else None, 'n': v['n']}
             for c, v in sorted(by_sup.items(), key=lambda kv: -kv[1]['kg'])]
    annual = []
    for y in sorted(by_year):
        v = by_year[y]
        ysup = [{'code': c, 'kg': round(d['kg'], 1), 'spend': round(d['spend'], 1),
                 'avg': round(d['spend'] / d['kg'], 3) if d['kg'] else None, 'n': d['n']}
                for (yy, c), d in sorted(by_ys.items(), key=lambda kv: -kv[1]['kg']) if yy == y]
        annual.append({'year': y, 'kg': round(v['kg'], 1), 'spend': round(v['spend'], 1),
                       'avg': round(v['spend'] / v['kg'], 3) if v['kg'] else None,
                       'min': round(v['min'], 3), 'max': round(v['max'], 3),
                       'n': v['n'], 'bySupplier': ysup})

    node = {'key': str(it['key']), 'name': it['name'], 'grp': it.get('grp'),
            'totalKg': round(tot_kg, 1), 'totalSpend': round(tot_spend, 1),
            'avgPrice': round(tot_spend / tot_kg, 3) if tot_kg else None,
            'minPrice': round(pmin, 3) if pmin is not None else None,
            'maxPrice': round(pmax, 3) if pmax is not None else None,
            'nTxn': len(purch),
            'firstMonth': monthly[0]['m'] if monthly else None,
            'lastMonth': monthly[-1]['m'] if monthly else None,
            'monthly': monthly, 'bySupplier': bysup, 'annual': annual}
    ret_rows = [{'veg': it['name'], 'date': _datestr(t['date']), 'code': t['code'],
                 'kg': round(t['inn'] or t['out'], 1), 'price': round(t['price'], 3)}
                for t in returns]
    return node, ret_rows


def integrity(it):
    """בדיקת שלמות: סכום כל הכניסות/יציאות (כל המסמכים) מול שורת הסה\"כ."""
    if it['total_in'] is None:
        return True, '— (אין שורת סה"כ)'
    s_in = sum(t['inn'] for t in it['txns'])
    s_out = sum(t['out'] for t in it['txns'])
    d_in = abs(s_in - it['total_in'])
    d_out = abs(s_out - (it['total_out'] or 0.0))
    ok = d_in < 0.5 and d_out < 0.5
    return ok, 'תקין' if ok else f'סטייה כניסה {d_in:.1f} / יציאה {d_out:.1f}'


def payload_from(nodes, rets, gen, extra=None):
    months = sorted({m['m'] for v in nodes for m in v['monthly'] if m['m']})
    years = sorted({a['year'] for v in nodes for a in v['annual']})
    sup = sorted({s['code'] for v in nodes for s in v['bySupplier']})
    p = {'generatedAt': gen, 'currency': '₪ לק"ג',
         'period': {'from': months[0] if months else None, 'to': months[-1] if months else None},
         'months': months, 'years': years,
         'suppliers': {c: SUPPLIER_LABELS.get(c, f'ספק {c}') for c in sup},
         'returns': {'totalKg': round(sum(r['kg'] for r in rets), 1), 'rows': rets},
         'totals': {'kg': round(sum(v['totalKg'] for v in nodes), 1),
                    'spend': round(sum(v['totalSpend'] for v in nodes), 1),
                    'txns': sum(v['nTxn'] for v in nodes), 'nVeg': len(nodes)}}
    if extra:
        p.update(extra)
    return p


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('inputs', nargs='+', help='כרטסות מלאי (xlsx רב-פריטים / פריט בודד / טקסט)')
    ap.add_argument('--out', default='veg_prices.json', help='פלט טאב הירקות')
    ap.add_argument('--mat-out', default=None, help='פלט טאב חומרי הגלם (כל הפריטים עם רכש)')
    ap.add_argument('--date', default=None, help="generatedAt (YYYY-MM-DD); ברירת מחדל: היום")
    args = ap.parse_args()

    paths = []
    for pat in args.inputs:
        paths.extend(sorted(glob.glob(pat)) or [pat])

    # פריט שמופיע ביותר מקובץ אחד -> **מיזוג תנועות** (כרטסת מלאה + כרטסות שבועיות).
    # התקופות לא אמורות לחפוף (מלאה עד חודש X, שבועיות אחריו) — אחרת כפילות.
    by_key = {}
    for p in paths:
        rdr = read_ledger_xlsx if p.lower().endswith(('.xlsx', '.xlsm')) else read_ledger_text
        for it in rdr(p):
            k = str(it['key'])
            if k in by_key:
                by_key[k]['txns'].extend(it['txns'])
                if it['total_in'] is not None:
                    by_key[k]['total_in'] = (by_key[k]['total_in'] or 0) + it['total_in']
                    by_key[k]['total_out'] = (by_key[k]['total_out'] or 0) + (it['total_out'] or 0)
            else:
                by_key[k] = dict(it, txns=list(it['txns']))
    print(f"נקראו {len(by_key)} פריטים מ-{len(paths)} קבצים")

    all_nodes, all_rets = [], []
    bad = 0
    for it in by_key.values():
        ok, msg = integrity(it)          # בדיקת שלמות מול הסה"כ ביחידות המקור
        if not ok:
            bad += 1
            print(f"⚠️  {it['name']} ({it['key']}): {msg}")
        nfix = _apply_unit_fixes(it)     # נרמול יחידות (אחרי הבדיקה, לפני האגרגציה)
        if nfix:
            print(f"🔧 {it['name']} ({it['key']}): נורמלו {nfix} תנועות לפי UNIT_FIXES")
        node, rets = build_item(it)
        if node['nTxn'] > 0 or rets:
            all_nodes.append(node)
            all_rets.extend(rets)
    gen = args.date or date.today().isoformat()

    # --- טאב הירקות: הטריים + משפחת העגבניות ---
    veg_nodes = []
    for n in all_nodes:
        if n['key'] in VEG_FRESH:
            veg_nodes.append(dict(n))
        elif n['key'] in TOMATO_FAMILY:
            fn = dict(n)
            fn['fam'] = TOMATO_FAMILY[n['key']]
            veg_nodes.append(fn)
    veg_nodes.sort(key=lambda v: (0 if v['key'] in VEG_FRESH else 1, -v['totalSpend']))
    veg_names = {v['name'] for v in veg_nodes}
    veg_rets = [r for r in all_rets if r['veg'] in veg_names]
    vp = payload_from(veg_nodes, veg_rets, gen, extra={'veg': veg_nodes})
    json.dump(vp, open(args.out, 'w', encoding='utf-8'), ensure_ascii=False, separators=(',', ':'))
    print(f"\nירקות ({args.out}): {len(veg_nodes)} פריטים | "
          f'{vp["totals"]["kg"]:,.0f} ק"ג | ₪{vp["totals"]["spend"]:,.0f} | החזרות: {len(veg_rets)}')
    for v in veg_nodes:
        tag = ' [משפחת עגבניות]' if v.get('fam') else ''
        print(f"   {v['name']:<24} {v['nTxn']:>4} קניות | {v['totalKg']:>11,.1f} ק\"ג | "
              f"ממוצע {v['avgPrice']}{tag}")

    # --- טאב חומרי הגלם: כל פריט עם רכש ---
    if args.mat_out:
        mat_nodes = sorted([n for n in all_nodes if n['nTxn'] > 0],
                           key=lambda v: -v['totalSpend'])
        groups = sorted({n['grp'] for n in mat_nodes if n['grp']})
        mp = payload_from(mat_nodes, all_rets, gen,
                          extra={'items': mat_nodes,
                                 'groups': {g: f'קבוצה {g}' for g in groups}})
        mp['totals']['nItems'] = len(mat_nodes)
        json.dump(mp, open(args.mat_out, 'w', encoding='utf-8'), ensure_ascii=False, separators=(',', ':'))
        print(f"\nחומרי גלם ({args.mat_out}): {len(mat_nodes)} פריטים | {len(groups)} קבוצות | "
              f'{mp["totals"]["kg"]:,.0f} יח\' | ₪{mp["totals"]["spend"]:,.0f} | '
              f"החזרות: {len(all_rets)} ({mp['returns']['totalKg']})")

    if bad:
        print(f"\n⚠️  {bad} פריטים נכשלו בבדיקת השלמות — לבדוק לפני קליטה.")
        sys.exit(2)


if __name__ == '__main__':
    main()
