#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ביקורת נתוני לקוחות מתוך קובץ הגלם של מחשבשבת (C_xxxx.xlsx, פורמט ציר+פירוט).

שני דברים:
1) תיקון סימן הכסף: בקובץ, אריזה חיובית (+) = זיכוי/החזרה, אריזה שלילית (−) = מכירה.
   חשבשבת לא תמיד הופכת את סימן הכסף בזיכוי, ולכן סכימה עיוורת מנפחת הכנסה.
   התיקון: כסף_מתוקן = +|כסף| במכירה, −|כסף| בזיכוי. אריזות כבר מתקזזות נכון.
2) מחיר לכל מכירה = |כסף|/|אריזות| בשורת מכירה, והרמת דגל למכירה חריגה —
   מחיר ששונה מהמחיר הרגיל של אותו לקוח לאותו פריט (טעות תמחור / מכירה מוזרה).

הרצה:
  python3 customer_audit.py <C_xxxx.xlsx> [--dev 3] [--csv out.csv]
"""
import argparse
import csv
from collections import defaultdict

import openpyxl

# עמודות בלשונית הפירוט (לפי מיקום קבוע בפורמט מחשבשבת החדש)
C_ACC, C_FAM, C_KEY, C_NAME, C_WEEK, C_DAY, C_BAL, C_KG, C_MONEY = 0, 1, 3, 4, 5, 6, 8, 9, 10


def detail_sheet(wb):
    for ws in wb.worksheets:
        hdr = " ".join(str(c.value or "") for c in next(ws.iter_rows(max_row=1)))
        if "שם חשבון" in hdr and "מפתח פריט" in hdr:
            return ws
    raise SystemExit("לא נמצאה לשונית פירוט לקוחות.")


def load_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = detail_sheet(wb)
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        acc, key, name, wk = r[C_ACC], r[C_KEY], r[C_NAME], r[C_WEEK]
        if acc is None or name is None or wk is None:
            continue
        bal = r[C_BAL] or 0
        money = r[C_MONEY] or 0
        rows.append(dict(acc=str(acc).strip(), key=str(key).strip(), name=str(name).strip(),
                         week=int(str(wk).strip()), bal=bal, money=money))
    return rows


def corrected_money(row):
    """כסף מתוקן: סימן נגזר מסימן האריזות (מכירה +, זיכוי −)."""
    if row["bal"] < 0:   # מכירה
        return abs(row["money"])
    if row["bal"] > 0:   # זיכוי
        return -abs(row["money"])
    return row["money"]


def normal_prices(rows):
    """מחיר 'רגיל' לכל (לקוח, פריט): המחיר השכיח (מעוגל לאגורה) על פני שורות המכירה,
    משוקלל באריזות. עמיד לחריגים."""
    buckets = defaultdict(lambda: defaultdict(float))  # (acc,name) -> price -> packages
    for r in rows:
        if r["bal"] >= 0:  # רק מכירות
            continue
        pk = -r["bal"]
        if pk <= 0:
            continue
        price = round(abs(r["money"]) / pk, 2)
        buckets[(r["acc"], r["name"])][price] += pk
    norm = {}
    for k, d in buckets.items():
        norm[k] = max(d.items(), key=lambda kv: kv[1])[0]  # המחיר עם הכי הרבה אריזות
    return norm


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--dev", type=float, default=3.0, help="סף סטייה באחוזים לדגל (ברירת מחדל 3%%)")
    ap.add_argument("--csv", help="נתיב לשמירת דוח המכירות החריגות")
    a = ap.parse_args()

    rows = load_rows(a.file)
    norm = normal_prices(rows)

    raw_total = sum(r["money"] for r in rows)
    corr_total = sum(corrected_money(r) for r in rows)
    n_credit = sum(1 for r in rows if r["bal"] > 0)
    n_sale = sum(1 for r in rows if r["bal"] < 0)

    # מצרפים מתוקנים פר לקוח
    by_cust = defaultdict(lambda: {"pkg": 0.0, "rev": 0.0})
    for r in rows:
        by_cust[r["acc"]]["pkg"] += -r["bal"]
        by_cust[r["acc"]]["rev"] += corrected_money(r)

    # זיהוי מכירות חריגות: שורת מכירה שמחירה סוטה מהמחיר הרגיל של אותו לקוח+פריט
    flags = []
    for r in rows:
        if r["bal"] >= 0:
            continue
        pk = -r["bal"]
        if pk <= 0:
            continue
        price = abs(r["money"]) / pk
        base = norm.get((r["acc"], r["name"]))
        if base is None or base == 0:
            continue
        dev = (price - base) / base * 100
        zero = (r["money"] == 0)
        if abs(dev) >= a.dev or zero:
            flags.append(dict(acc=r["acc"], name=r["name"], week=r["week"], pkg=pk,
                              money=r["money"], price=round(price, 2), normal=base,
                              dev=round(dev, 1), impact=abs(r["money"] - base * pk)))

    flags.sort(key=lambda f: -f["impact"])

    print("=== תיקון סימן הכסף ===")
    print(f"שורות: {len(rows):,} | מכירות {n_sale:,} | זיכויים {n_credit:,}")
    print(f"הכנסה כפי שמחושבת היום (סכום עיוור): ₪{raw_total:,.0f}")
    print(f"הכנסה מתוקנת (זיכוי מוריד):          ₪{corr_total:,.0f}")
    print(f"הפרש (ניפוח מזיכויים):               ₪{raw_total - corr_total:,.0f} "
          f"({(raw_total - corr_total) / corr_total * 100:.2f}%)")

    print(f"\n=== מכירות חריגות (סטייה ≥ {a.dev:g}% מהמחיר הרגיל של אותו לקוח+פריט) ===")
    print(f"נמצאו {len(flags)} מכירות חריגות. עשרים המשמעותיות ביותר:")
    print(f"{'לקוח':22} {'פריט':22} {'שב':>3} {'אריז':>5} {'מחיר':>7} {'רגיל':>7} {'סטייה%':>7} {'השפעה₪':>9}")
    for f in flags[:20]:
        print(f"{f['acc'][:22]:22} {f['name'][:22]:22} {f['week']:>3} {f['pkg']:>5.0f} "
              f"{f['price']:>7.2f} {f['normal']:>7.2f} {f['dev']:>7.1f} {f['impact']:>9,.0f}")

    if a.csv:
        with open(a.csv, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.DictWriter(fh, fieldnames=["acc", "name", "week", "pkg", "money",
                                               "price", "normal", "dev", "impact"])
            w.writeheader()
            w.writerows(flags)
        print(f"\nדוח מלא נשמר: {a.csv} ({len(flags)} שורות)")


if __name__ == "__main__":
    main()
