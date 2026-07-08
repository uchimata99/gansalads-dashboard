#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_purchasing.py — בונה את קטלוג הרכש של purchasing.html מארבעה קבצי מחשבשבת.

מקורות (xlsx שיוצאו ממחשבשבת, נשמרים בדרייב בתיקיית "אפליקציית רכש"):
  --items     אינדקס פריטים            (קוד מיון, מפתח פריט, שם פריט)  = הפריטים הרלוונטיים
  --suppliers אינדקס ספקים             (קוד מיון, מפתח חשבון, שם החשבון)
  --prices    מחיר קנייה אחרון חומרי גלם(מפתח פריט, תאריך קניה אחרון, מחיר קניה אחרון)
  --map       שיוך ספק לחומר גלם        (מפתח חשבון, מפתח פריט, שם פריט)

הפלט: purchasing_catalog.json במבנה שהדף מצפה לו:
  { generatedAt, cats:{code:label}, suppliers:{acctKey:name},
    items:[ {k,n,c,p,pd,s:[supplierKeys]} ] }

כללי ניקוי:
  * הפריטים הרלוונטיים = אלה שבאינדקס הפריטים (dedupe לפי מפתח).
  * שיוך פריט→ספק נלקח מקובץ השיוך, אבל **רק** מפתחות חשבון שקיימים באינדקס
    הספקים (מסנן החוצה לקוחות 10xxx ו"ללא" שמופיעים בקובץ תנועות הגלם).
  * מחיר 0 או תאריך 01/01/80 = "אין רכש אמיתי" ⇒ מחיר null (לא ממציאים מספר).
  * מפתחות מנורמלים למחרוזת (מסיר .0 מצף).

הרצה:
  python3 build_purchasing.py --items items_index.xlsx --suppliers suppliers_index.xlsx \
      --prices last_price.xlsx --map supplier_item_map.xlsx --out purchasing_catalog.json
בדיקת הצטלבות פנימית מודפסת בסוף (מספר פריטים, כמה עם ספק, כמה עם מחיר).
"""
import argparse, json, datetime, re
import openpyxl

# תוויות תצוגה לקוד המיון (עזר תצוגה בלבד — הקוד המספרי תמיד מוצג לצדן)
CAT_LABELS = {
    "200": "חומרי גלם / קניות",
    "201": "אריזות",
    "202": "ביצים",
    "203": "שימורים וירקות מעובדים",
    "204": "תבלינים",
    "205": "קטניות, אגוזים ודגנים",
    "206": "ירקות ופירות טריים",
    "207": 'ירקות מעובדים (בד"ץ)',
    "208": "ניקיון וחד-פעמי",
    "210": "שמנים, רטבים ומיצים",
    "211": "כימיקלים ותוספים",
    "221": "מגוון (דגים, גבינות, פסטה)",
}


def norm_key(v):
    """מפתח -> מחרוזת נקייה (מסיר .0 מצף, רווחים)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def clean_name(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def find_header(ws, wanted, max_scan=8):
    """מאתר את שורת הכותרת ומחזיר {שם לוגי: index עמודה}. wanted: {logical:[substrings]}."""
    for r_i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        cols = {}
        for c_i, val in enumerate(row):
            if val is None:
                continue
            s = str(val).strip()
            for logical, subs in wanted.items():
                if logical in cols:
                    continue
                for sub in subs:
                    if sub in s:
                        cols[logical] = c_i
                        break
        if len(cols) == len(wanted):
            return r_i + 1, cols
    raise SystemExit("לא נמצאה שורת כותרת עם כל השדות: %s (נמצא: %s)" % (list(wanted), cols))


def load_items(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr_row, cols = find_header(ws, {
        "code": ["קוד מיון"],
        "key":  ["מפתח פריט"],
        "name": ["שם פריט"],
    })
    ci, ki, ni = cols["code"], cols["key"], cols["name"]
    items, seen = [], set()
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        code = row[ci] if ci < len(row) else None
        key = row[ki] if ki < len(row) else None
        name = row[ni] if ni < len(row) else None
        # שורות מפריד ("קוד מיון",201,..) או ריקות — קוד לא מספרי או חסר מפתח/שם
        code_s = norm_key(code)
        if not re.fullmatch(r"\d{3}", code_s):
            continue
        k, n = norm_key(key), clean_name(name)
        # טיפול בשורה הפוכה (שם במקום מפתח): אם k לא מספרי אך n כן — החלפה
        if not re.fullmatch(r"\d+", k) and re.fullmatch(r"\d+", n.replace(" ", "")):
            k, n = norm_key(name), clean_name(key)
        if not re.fullmatch(r"\d+", k) or not n:
            continue
        if k in seen:
            continue
        seen.add(k)
        items.append({"k": k, "n": n, "c": code_s})
    return items


def load_suppliers(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr_row, cols = find_header(ws, {
        "key":  ["מפתח חשבון"],
        "name": ["שם החשבון", "שם חשבון"],
    })
    ki, ni = cols["key"], cols["name"]
    sup = {}
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        key = row[ki] if ki < len(row) else None
        name = row[ni] if ni < len(row) else None
        k, n = norm_key(key), clean_name(name)
        if re.fullmatch(r"\d+", k) and n:
            sup[k] = n
    return sup


def load_prices(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr_row, cols = find_header(ws, {
        "key":   ["מפתח פריט"],
        "date":  ["תאריך קניה אחרון", "תאריך קנייה אחרון"],
        "price": ["מחיר קניה אחרון", "מחיר קנייה אחרון"],
    })
    ki, di, pi = cols["key"], cols["date"], cols["price"]
    prices = {}
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        k = norm_key(row[ki] if ki < len(row) else None)
        if not re.fullmatch(r"\d+", k):
            continue
        raw_p = row[pi] if pi < len(row) else None
        raw_d = row[di] if di < len(row) else None
        # מחיר -> float
        p = None
        if raw_p is not None and str(raw_p).strip() != "":
            try:
                p = float(str(raw_p).replace(",", ""))
            except ValueError:
                p = None
        # תאריך -> dd/mm/yy
        d = ""
        if isinstance(raw_d, (datetime.datetime, datetime.date)):
            d = raw_d.strftime("%d/%m/%y")
        elif raw_d is not None:
            d = str(raw_d).strip()
        # 0 או תאריך 1980 = אין רכש אמיתי
        is_dummy = (d.startswith("01/01/80") or d.startswith("1/1/80"))
        if p is None or p == 0 or is_dummy:
            prices[k] = (None, "")
        else:
            prices[k] = (round(p, 2), d)
    return prices


def load_map(path, valid_suppliers):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # לשונית הנתונים (לא "Claude Log")
    ws = None
    for name in wb.sheetnames:
        if "Claude Log" not in name:
            ws = wb[name]
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    hdr_row, cols = find_header(ws, {
        "acct": ["מפתח חשבון"],
        "item": ["מפתח פריט"],
    })
    ai, ii = cols["acct"], cols["item"]
    m = {}
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        a = norm_key(row[ai] if ai < len(row) else None)
        it = norm_key(row[ii] if ii < len(row) else None)
        if not re.fullmatch(r"\d+", it):
            continue
        if a not in valid_suppliers:      # רק ספקים אמיתיים (מסנן לקוחות/ללא)
            continue
        m.setdefault(it, [])
        if a not in m[it]:
            m[it].append(a)
    return m


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--items", required=True)
    ap.add_argument("--suppliers", required=True)
    ap.add_argument("--prices", required=True)
    ap.add_argument("--map", required=True)
    ap.add_argument("--out", default="purchasing_catalog.json")
    ap.add_argument("--date", default=None, help="חותמת generatedAt (dd/mm/yyyy); ברירת מחדל=היום")
    a = ap.parse_args()

    items = load_items(a.items)
    suppliers = load_suppliers(a.suppliers)
    prices = load_prices(a.prices)
    item_sup = load_map(a.map, set(suppliers.keys()))

    used_sup = set()
    with_sup = with_price = 0
    for it in items:
        s = item_sup.get(it["k"], [])
        # מיון ספקי הפריט לפי שם לתצוגה יציבה
        s = sorted(s, key=lambda k: suppliers.get(k, k))
        it["s"] = s
        used_sup.update(s)
        p, pd = prices.get(it["k"], (None, ""))
        it["p"] = p
        it["pd"] = pd
        if s:
            with_sup += 1
        if p is not None:
            with_price += 1

    # רק ספקים שבשימוש בפועל (מקטין את המטען) — אך שומר את השם המלא
    suppliers_out = {k: suppliers[k] for k in used_sup if k in suppliers}

    cats = {}
    for it in items:
        c = it["c"]
        cats[c] = CAT_LABELS.get(c, "קוד " + c)

    gen = a.date or datetime.date.today().strftime("%d/%m/%Y")
    out = {
        "generatedAt": gen,
        "cats": cats,
        "suppliers": suppliers_out,
        "items": items,
    }
    with open(a.out, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))

    print("== קטלוג רכש נבנה ==")
    print("פריטים (רלוונטיים):", len(items))
    print("קטגוריות:", ", ".join("%s=%s" % (c, cats[c]) for c in sorted(cats)))
    print("ספקים בקובץ האינדקס:", len(suppliers), "· בשימוש בקטלוג:", len(suppliers_out))
    print("פריטים עם ספק משויך:", with_sup, "· ללא ספק:", len(items) - with_sup)
    print("פריטים עם מחיר קנייה אחרון:", with_price, "· ללא מחיר:", len(items) - with_price)
    print("נכתב אל:", a.out)


if __name__ == "__main__":
    main()
