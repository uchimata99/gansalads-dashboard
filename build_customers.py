#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
בונה לקוחות קנוני — מייצר את אובייקט D של דשבורד הלקוחות מקובץ הגלם C_xxxx.xlsx.

זה מוציא סופית את בניית D מהצ'אט אל קוד מחויב. כל כללי התיקון בפנים:
  • תיקון סימן הכסף (מ-customer_audit): אריזה −(מכירה) ⇒ +|כסף| ; אריזה +(זיכוי) ⇒ −|כסף|.
  • דגלים אדומים: מכירות חריגות (מחיר שסוטה מהמחיר הרגיל של אותו לקוח+פריט / מחיר 0).

הפלט הוא קובץ JSON של D, במבנה המדויק ש-dashboard_customers.html מצפה לו. אחר כך
ingest_customers.py מעלה אותו ללשונית CUSTOMERS בגיליון המוגבל (הדף לא משתנה).

מבנה D (חוזה הדף): meta, customers[], weekly{}, distribution, price_prods[],
price_gaps[], item_kg{}, items[], fam_money{}, coverage{}, anom_totals,
insights, red_flags[], gp_meta(=null עד שיהיו עלויות).

הרצה
----
  python3 build_customers.py C_1_26.xlsx                         # יבש: סיכום + בדיקות
  python3 build_customers.py C_1_26.xlsx --out customers_D.json  # כותב את D
  python3 build_customers.py C_1_26.xlsx --out D.json --production ../gansalads---engine/seed_payload.json
       # ממלא גם את צד הייצור בהצלבת הכיסוי (coverage.prod) מסיכומי הייצור
  python3 build_customers.py C_1_26.xlsx --flags-csv red_flags.csv
"""
import argparse
import csv
import json
import re
from collections import defaultdict
from statistics import mean, median, pstdev

import openpyxl

# עמודות בלשונית הפירוט (פורמט מחשבשבת החדש; זהה ל-customer_audit)
C_ACC, C_FAM, C_KEY, C_NAME, C_WEEK, C_DAY, C_BAL, C_KG, C_MONEY = 0, 1, 3, 4, 5, 6, 8, 9, 10

DEV_FLAG = 3.0   # סף סטייה באחוזים לדגל מכירה חריגה
RECENT_W = 8     # חלון שבועות אחרונים להצגת דגלים (הדגלים נועדו לפעולה, לא להיסטוריה)


def detail_sheet(wb):
    for ws in wb.worksheets:
        hdr = " ".join(str(c.value or "") for c in next(ws.iter_rows(max_row=1)))
        if "שם חשבון" in hdr and "מפתח פריט" in hdr:
            return ws
    raise SystemExit("לא נמצאה לשונית פירוט לקוחות בקובץ.")


def load_rows(path):
    """שורות גלם עשירות: לקוח, משפחה, מפתח, פריט, שבוע, אריזות(bal), משקל, כסף."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = detail_sheet(wb)
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        acc, name, wk = r[C_ACC], r[C_NAME], r[C_WEEK]
        if acc is None or name is None or wk is None:
            continue
        try:
            week = int(str(wk).strip())
        except (TypeError, ValueError):
            continue
        rows.append(dict(
            acc=str(acc).strip(),
            fam=str(r[C_FAM] or "").strip() or "אחר",
            key=str(r[C_KEY] or "").strip(),
            name=str(name).strip(),
            week=week,
            bal=r[C_BAL] or 0,        # אריזות: שלילי=מכירה, חיובי=זיכוי
            kg=r[C_KG] or 0,          # משקל כולל (מגניטודה)
            money=r[C_MONEY] or 0,
        ))
    if not rows:
        raise SystemExit("הקובץ נטען אך לא נמצאו שורות פירוט תקינות.")
    return rows


def parse_weight(name):
    """משקל אריזה (ק\"ג) נגזר משם הפריט — זהה לכלל המנוע. None אם לא ניתן לגזור."""
    s = (name or "").strip()
    m = re.search(r'(\d+(?:\.\d+)?)\s*(גרם|גר|ליטר|ק"?ג|קג|ק"כ|ל)', s)
    if m:
        n = float(m.group(1)); u = m.group(2)
        return n / 1000 if u in ("גרם", "גר") else n
    m2 = re.search(r"(\d+(?:\.\d+)?)\s*$", s)        # מספר בודד בסוף (למשל 'משוויאה 250')
    if m2:
        n = float(m2.group(1))
        return n / 1000 if n >= 50 else n
    return None


def dirs(row):
    """אריזות וכסף מכווננים: מכירה חיובית, זיכוי שלילי.
    הערה חשובה: עמודת המשקל בקובץ הגלם **אינה אמינה** — בזיכוי המשקל מופיע בפלוס
    (לא במינוס), כך שסכימה עיוורת מנפחת (אפשר אפס אריזות אך משקל חיובי). לכן הק\"ג
    מחושב תמיד = אריזות נטו × משקל אריזה (parse_weight), לעולם לא מהעמודה הגולמית."""
    if row["bal"] < 0:        # מכירה
        return -row["bal"], abs(row["money"])
    if row["bal"] > 0:        # זיכוי
        return -row["bal"], -abs(row["money"])
    return 0.0, row["money"]


def sd(vals):
    return round(pstdev(vals), 2) if len(vals) > 1 else 0.0


def build_D(rows, prod_rev=None):
    weeks = sorted({r["week"] for r in rows})
    wstr = [str(w) for w in weeks]

    # ── צבירה פר לקוח, פר (לקוח,פריט), פר (לקוח,שבוע,פריט), פר פריט, פר משפחה ──
    cust = defaultdict(lambda: dict(money=0.0, qty=0.0, kg=0.0, weeks=set(),
                                    by_week=defaultdict(float), by_fam=defaultdict(float)))
    cust_item = defaultdict(lambda: dict(money=0.0, qty=0.0, kg=0.0))     # (acc,item)
    cwi = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))            # (acc,week)->item->[money,qty]
    item_tot = defaultdict(lambda: dict(money=0.0, qty=0.0, kg=0.0, fam="אחר"))
    fam_money = defaultdict(float)
    buyer = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))          # item->acc->[money,qty] (מכירות)

    for r in rows:
        pk, mon = dirs(r)
        a, it, fam, w = r["acc"], r["name"], r["fam"], r["week"]
        kg = pk * (parse_weight(it) or 0.0)   # ק"ג = אריזות נטו × משקל אריזה (לא מהעמודה הגולמית)
        c = cust[a]
        c["money"] += mon; c["qty"] += pk; c["kg"] += kg
        c["weeks"].add(w); c["by_week"][w] += mon; c["by_fam"][fam] += mon
        ci = cust_item[(a, it)]; ci["money"] += mon; ci["qty"] += pk; ci["kg"] += kg
        cwi[(a, w)][it][0] += mon; cwi[(a, w)][it][1] += pk
        t = item_tot[it]; t["money"] += mon; t["qty"] += pk; t["kg"] += kg; t["fam"] = fam
        fam_money[fam] += mon
        if r["bal"] < 0:                      # מכירות בלבד למחירים
            b = buyer[it][a]; b[0] += abs(r["money"]); b[1] += -r["bal"]

    # ── רשימת מוצרים (מאונדקסת לפי כסף יורד) + משקל לאריזה ──
    items_sorted = sorted(item_tot, key=lambda it: -item_tot[it]["money"])
    item_idx = {it: i for i, it in enumerate(items_sorted)}
    item_kg = {}
    for it in item_tot:
        w = parse_weight(it)
        if w:
            item_kg[it] = round(w, 3)   # משקל אריזה ישיר מהשם — אמין, לא מהעמודה הגולמית

    # ── רשימת לקוחות (מדורגת לפי כסף יורד) ──
    accs = sorted(cust, key=lambda a: -cust[a]["money"])
    acc_idx = {a: i for i, a in enumerate(accs)}
    customers = []
    for a in accs:
        c = cust[a]
        nw = len(c["weeks"])
        wk_vals = [c["by_week"][w] for w in sorted(c["weeks"]) if w != 1]  # שבוע 1 חלקי — מחוץ לממוצע
        # מוצרים מובילים ללקוח
        mine = [(it, cust_item[(a, it)]) for (aa, it) in cust_item if aa == a]
        mine.sort(key=lambda kv: -kv[1]["money"])
        top_prod = []
        for it, d in mine[:12]:
            top_prod.append(dict(item=it, money=round(d["money"], 2), qty=round(d["qty"], 1),
                                 cost=None, profit=None, profit_kg=None))
        avg_box = round(c["money"] / c["qty"], 2) if c["qty"] else 0.0
        avg_kg = round(c["money"] / c["kg"], 2) if c["kg"] else None
        customers.append(dict(
            name=a, money=round(c["money"], 2), qty=round(c["qty"], 1),
            wk_mean=round(mean(wk_vals), 2) if wk_vals else 0.0, wk_sd=sd(wk_vals),
            n_weeks=nw, avg_price=avg_box, avg_box=avg_box, avg_kg=avg_kg,
            n_prod=len(mine), top_prod=top_prod,
            by_week={str(w): round(c["by_week"][w], 2) for w in sorted(c["weeks"])},
            by_fam={f: round(v, 2) for f, v in sorted(c["by_fam"].items(), key=lambda kv: -kv[1])},
            gp=None))

    # ── weekly: week -> custIdx(str) -> {m,q,p:[[itemIdx,money,qty]]} ──
    weekly = {}
    for (a, w), items in cwi.items():
        ws_ = weekly.setdefault(str(w), {})
        p = sorted(([item_idx[it], round(mo, 2), round(q, 1)] for it, (mo, q) in items.items()),
                   key=lambda x: -x[1])
        m = round(sum(x[1] for x in p), 2)
        q = round(sum(x[2] for x in p), 1)
        ws_[str(acc_idx[a])] = dict(m=m, q=q, p=p)

    # ── price_prods + price_gaps ──
    price_prods, price_gaps = [], []
    for it in items_sorted:
        bl = buyer.get(it, {})
        rows_b = []
        for a, (mo, q) in bl.items():
            if q <= 0:
                continue
            rows_b.append(dict(name=a, price=round(mo / q, 2), qty=round(q, 1),
                               money=round(mo, 2), low=(q < 5)))
        if not rows_b:
            continue
        tot_m = sum(b["money"] for b in rows_b)
        tot_q = sum(b["qty"] for b in rows_b)
        avg = round(tot_m / tot_q, 2) if tot_q else 0.0
        prices = [b["price"] for b in rows_b]
        for b in rows_b:
            b["diff_pct"] = round((b["price"] - avg) / avg * 100, 1) if avg else 0.0
        rows_b.sort(key=lambda b: -b["qty"])
        price_prods.append(dict(item=it, avg=avg, price_sd=sd(prices),
                                min=round(min(prices), 2), max=round(max(prices), 2),
                                n_buyers=len(rows_b), buyers=rows_b))
        # פערי מחיר — לפי קונים אמינים (qty>=5)
        core = [b["price"] for b in rows_b if not b["low"]]
        if len(core) >= 3:
            med = median(core)
            spread = round((max(core) - min(core)) / med * 100, 1) if med else 0.0
            if spread >= 15:
                price_gaps.append(dict(item=it, fam=item_tot[it]["fam"], avg=avg,
                                       core_sd=sd(core), core_min=round(min(core), 2),
                                       core_max=round(max(core), 2), core_med=round(med, 2),
                                       core_spread_pct=spread, n_buyers=len(core)))
    price_gaps.sort(key=lambda g: -g["core_spread_pct"])

    # ── distribution ──
    monies = [c["money"] for c in customers]
    total_money = round(sum(monies), 2)
    total_qty = round(sum(c["qty"] for c in customers), 1)
    bands = [("₪0–25K", 0, 25000), ("₪25–50K", 25000, 50000), ("₪50–100K", 50000, 100000),
             ("₪100–250K", 100000, 250000), ("₪250K+", 250000, float("inf"))]
    segments = []
    for label, lo, hi in bands:
        grp = [m for m in monies if lo <= m < hi]
        if grp:
            segments.append(dict(band=label, n=len(grp), money=round(sum(grp), 2)))
    pareto, cum = [], 0.0
    for i, m in enumerate(sorted(monies, reverse=True), 1):
        cum += m
        pareto.append(dict(rank=i, cum_pct=round(cum / total_money * 100, 1) if total_money else 0.0))
    whist = defaultdict(int)
    for c in customers:
        whist[c["n_weeks"]] += 1
    dist_weeks = [dict(weeks=w, n=n) for w, n in sorted(whist.items())]
    distribution = dict(cust_mean=round(mean(monies), 2) if monies else 0.0,
                        cust_sd=sd(monies), segments=segments, pareto=pareto, weeks=dist_weeks)

    # ── meta ── (שבוע 1 חלקי — מחוץ לממוצע השבועי)
    wk_totals = [sum(c["by_week"].get(w, 0.0) for c in cust.values()) for w in weeks if w != 1]
    wk_mean = round(mean(wk_totals), 2) if wk_totals else 0.0
    wk_sd = sd(wk_totals)
    box_prices = [c["avg_box"] for c in customers if c["avg_box"]]
    meta = dict(n_customers=len(customers), weeks=weeks, total_money=total_money,
                total_qty=total_qty, wk_mean=wk_mean, wk_sd=wk_sd,
                wk_cv=str(round(wk_sd / wk_mean * 100)) if wk_mean else "0",
                price_sd=sd(box_prices))

    # ── coverage + anom_totals (הצלבה מול ייצור, אם סופק) ──
    coverage = {}
    cov_pcts = []
    for w in weeks:
        cu = round(sum(c["by_week"].get(w, 0.0) for c in cust.values()), 2)
        pr = round(prod_rev.get(w), 2) if (prod_rev and prod_rev.get(w)) else None
        coverage[str(w)] = dict(cust=cu, prod=pr)
        if pr:
            cov_pcts.append(cu / pr * 100)
    anom_totals = dict(avg_cov=round(mean(cov_pcts)) if cov_pcts else None)

    # ── insights (תובנות שבועיות) + red_flags (דגלים אדומים) ──
    insights = build_insights(rows, weeks, cust, cust_item)
    red_flags = build_red_flags(rows)

    return dict(meta=meta, customers=customers, weekly=weekly, distribution=distribution,
                price_prods=price_prods, price_gaps=price_gaps, item_kg=item_kg,
                items=items_sorted, fam_money={f: round(v, 2) for f, v in fam_money.items()},
                coverage=coverage, anom_totals=anom_totals, insights=insights,
                red_flags=red_flags, gp_meta=None)


def build_insights(rows, weeks, cust, cust_item):
    """תובנות שבוע אחרון מול ממוצע עד 6 שבועות קודמים: לקוחות שנעלמו/צנחו/זינקו ומוצרים."""
    if len(weeks) < 2:
        return None
    cur = weeks[-1]
    base_ws = weeks[-7:-1] if len(weeks) > 1 else []
    base_lbl = f"{base_ws[0]}–{base_ws[-1]}" if base_ws else ""
    # כסף פר (לקוח,שבוע) ופר (לקוח,פריט,שבוע)
    cw = defaultdict(lambda: defaultdict(float))
    cwiq = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))  # (acc,item)->week->[money,qty]
    for r in rows:
        _, mon = dirs(r)
        pk = -r["bal"] if r["bal"] < 0 else 0.0
        cw[r["acc"]][r["week"]] += mon
        cwiq[(r["acc"], r["name"])][r["week"]][0] += mon
        cwiq[(r["acc"], r["name"])][r["week"]][1] += pk

    def avg_base(d):
        vals = [d[w] for w in base_ws if w in d]
        return mean(vals) if vals else 0.0

    gone, drop, surge = [], [], []
    for a, d in cw.items():
        ab = avg_base(d)
        cur_m = d.get(cur, 0.0)
        recN = sum(1 for w in base_ws if d.get(w, 0) > 0)
        if ab > 0 and cur_m == 0 and recN >= 2:
            gone.append(dict(name=a, avg=round(ab), recN=recN))
        elif ab > 0 and cur_m > 0:
            pct = round((cur_m - ab) / ab * 100)
            if pct <= -40:
                drop.append(dict(name=a, w26=round(cur_m), avg=round(ab), pct=abs(pct)))
            elif pct >= 80:
                surge.append(dict(name=a, w26=round(cur_m), avg=round(ab), pct=pct))
    gone.sort(key=lambda x: -x["avg"]); drop.sort(key=lambda x: -x["avg"]); surge.sort(key=lambda x: -x["w26"])

    pdrop, psurge, pnew = [], [], []
    for (a, it), d in cwiq.items():
        bm = [d[w][0] for w in base_ws if w in d]
        bq = [d[w][1] for w in base_ws if w in d]
        avgM = mean(bm) if bm else 0.0
        avgQ = mean(bq) if bq else 0.0
        m26, q26 = d.get(cur, [0.0, 0.0])
        if avgM > 50 and m26 == 0 and len([1 for w in base_ws if w in d]) >= 2:
            pdrop.append(dict(name=a, item=it, avgQ=round(avgQ, 1), avgM=round(avgM)))
        elif avgM > 0 and m26 > 0 and avgM and m26 / avgM >= 2.5 and m26 > 200:
            psurge.append(dict(name=a, item=it, m26=round(m26), q26=round(q26, 1),
                               avgM=round(avgM), ratio=round(m26 / avgM, 1)))
        elif avgM == 0 and m26 > 200:
            pnew.append(dict(name=a, item=it, m26=round(m26), q26=round(q26, 1)))
    pdrop.sort(key=lambda x: -x["avgM"]); psurge.sort(key=lambda x: -x["m26"]); pnew.sort(key=lambda x: -x["m26"])

    flags = []
    tot_cur = sum(d.get(cur, 0.0) for d in cw.values())
    tot_base = mean([sum(d.get(w, 0.0) for d in cw.values()) for w in base_ws]) if base_ws else 0
    if tot_base and (tot_cur - tot_base) / tot_base <= -0.30:
        flags.append(f"סך ההזמנות השבוע נמוך ב-{round((1-tot_cur/tot_base)*100)}% מהממוצע")

    return dict(week=cur, base=base_lbl, flags=flags or None,
                gone=gone[:15] or None, drop=drop[:15] or None, surge=surge[:15] or None,
                pdrop=pdrop[:15] or None, psurge=psurge[:15] or None, pnew=pnew[:15] or None)


def build_red_flags(rows):
    """דגלים אדומים: מכירות *חריגות באמת* — מחיר שסוטה ≥DEV_FLAG% מהצפוי, או מחיר 0.

    'הצפוי' = המחיר האחרון של אותו לקוח+פריט, *מתוקן לפי השינוי הכללי של אותו שבוע*.
    כשמיכל מעלה מחירון (עדכון רוחבי של 4–6% לכולם), חציון השינוי השבועי סופג את
    העלייה, כך שלקוח שקיבל בדיוק את העלייה הכללית אינו מסומן — מסומן רק מי שחורג
    מעבר לה (למשל +11% כשכולם קיבלו +5%), או מחיר 0. מוצגים רק RECENT_W השבועות
    האחרונים — הכלי נועד לפעולה על טעויות תמחור עכשוויות, לא לסקירת היסטוריה."""
    sales = defaultdict(list)   # (לקוח,פריט) -> [(שבוע, מחיר, אריזות, כסף, שורה)]
    for r in rows:
        if r["bal"] >= 0:
            continue
        pk = -r["bal"]
        if pk <= 0:
            continue
        sales[(r["acc"], r["name"])].append(
            (r["week"], abs(r["money"]) / pk, pk, r["money"], r))
    if not sales:
        return []

    # שלב 1: לכל מכירה — בסיס = מחיר מודלי של שבוע-המכירה הקודם, וסטייה גולמית מולו
    entries = []   # dict לכל מכירה עם dev גולמי (None כשאין מחיר קודם)
    for lst in sales.values():
        weeks_sorted = sorted({s[0] for s in lst})
        for (wk, price, pk, money, r) in lst:
            prev = [w for w in weeks_sorted if w < wk]
            base = 0.0
            if prev:
                buck = defaultdict(float)
                for (w2, p2, pk2, m2, r2) in lst:
                    if w2 == prev[-1]:
                        buck[round(p2, 2)] += pk2
                if buck:
                    base = max(buck.items(), key=lambda kv: kv[1])[0]
            dev = (price - base) / base * 100 if (base > 0 and money != 0) else None
            entries.append(dict(wk=wk, cust=r["acc"], item=r["name"], pk=pk, price=price,
                                money=money, base=base, dev=dev))

    # שלב 2: לכל שבוע, "הצפוי" = {0 (ללא שינוי), עוצמת העדכון הרוחבי}.
    # עוצמת העדכון = חציון הסטיות החיוביות המהותיות (≥DEV_FLAG) של אותו שבוע.
    # פריסה מדורגת של עליית מחיר יוצרת פיזור דו-שיאי (0 ו-+5%) — שניהם לגיטימיים.
    week_bump = {}
    per_week = defaultdict(list)
    for e in entries:
        if e["dev"] is not None:
            per_week[e["wk"]].append(e["dev"])
    for w, v in per_week.items():
        pos = [d for d in v if d >= DEV_FLAG]
        week_bump[w] = median(pos) if len(pos) >= max(5, 0.05 * len(v)) else 0.0

    # שלב 3: דגל רק אם הסטייה רחוקה מ*כל* ערך צפוי (ירידת מחיר, זינוק חריג), או מחיר 0
    max_week = max(r["week"] for r in rows)
    cutoff = max_week - RECENT_W + 1
    flags = []
    for e in entries:
        if e["wk"] < cutoff:
            continue
        if e["money"] == 0:
            flags.append(dict(cust=e["cust"], item=e["item"], week=e["wk"], pkg=round(e["pk"], 1),
                              price=0.0, normal=round(e["base"], 2),
                              dev=(-100.0 if e["base"] else 0.0),
                              impact=round(e["base"] * e["pk"], 2), zero=True))
            continue
        if e["dev"] is None or e["base"] <= 0:
            continue   # אין מחיר קודם להשוות מולו — לא ניתן לשפוט
        bump = week_bump.get(e["wk"], 0.0)
        expected = [0.0] + ([bump] if bump else [])
        near = min(expected, key=lambda x: abs(e["dev"] - x))
        resid = e["dev"] - near
        if abs(resid) >= DEV_FLAG:
            norm = e["base"] * (1 + near / 100)
            flags.append(dict(cust=e["cust"], item=e["item"], week=e["wk"], pkg=round(e["pk"], 1),
                              price=round(e["price"], 2), normal=round(norm, 2), dev=round(resid, 1),
                              impact=round(abs(e["money"] - norm * e["pk"]), 2), zero=False))
    flags.sort(key=lambda f: -f["impact"])
    return flags


def load_prod_rev(path):
    """מפת שבוע→הכנסת ייצור מסיכומי המנוע (seed_payload/DATA), להצלבת כיסוי."""
    p = json.load(open(path, encoding="utf-8"))
    return {w["w"]: w.get("rev", 0) for w in p.get("WEEKS", [])}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("file", help="קובץ הגלם C_xxxx.xlsx")
    ap.add_argument("--out", help="נתיב לכתיבת אובייקט D (JSON)")
    ap.add_argument("--production", help="seed_payload.json של הייצור (להצלבת כיסוי)")
    ap.add_argument("--flags-csv", help="נתיב לשמירת הדגלים האדומים כ-CSV")
    a = ap.parse_args()

    rows = load_rows(a.file)
    prod_rev = load_prod_rev(a.production) if a.production else None
    D = build_D(rows, prod_rev)

    raw = sum(r["money"] for r in rows)
    n_sale = sum(1 for r in rows if r["bal"] < 0)
    n_cred = sum(1 for r in rows if r["bal"] > 0)
    print("=== בונה לקוחות ===")
    print(f"שורות: {len(rows):,} | מכירות {n_sale:,} | זיכויים {n_cred:,}")
    print(f"לקוחות: {D['meta']['n_customers']} | מוצרים: {len(D['items'])} | שבועות: {D['meta']['weeks'][0]}–{D['meta']['weeks'][-1]}")
    print(f"הכנסה מתוקנת: ₪{D['meta']['total_money']:,.0f}  (סכום עיוור: ₪{raw:,.0f} ; ניפוח מזיכויים ₪{raw-D['meta']['total_money']:,.0f})")
    print(f"אריזות (נטו): {D['meta']['total_qty']:,.0f} | ממוצע שבועי ₪{D['meta']['wk_mean']:,.0f} (CV {D['meta']['wk_cv']}%)")
    print(f"דגלים אדומים: {len(D['red_flags'])} מכירות חריגות")
    if D["insights"]:
        ins = D["insights"]
        print(f"תובנות שבוע {ins['week']} (מול {ins['base']}): "
              f"נעלמו {len(ins['gone'] or [])}, צנחו {len(ins['drop'] or [])}, זינקו {len(ins['surge'] or [])}")
    # בדיקת שפיות: סכום by_week של כל הלקוחות = total_money
    chk = round(sum(sum(c["by_week"].values()) for c in D["customers"]), 2)
    if abs(chk - D["meta"]["total_money"]) > 1:
        raise SystemExit(f"בדיקת הצטלבות נכשלה: by_week={chk:,.0f} ≠ total={D['meta']['total_money']:,.0f}")
    print(f"בדיקת הצטלבות פנימית: ✓ (סכום by_week = total_money = ₪{chk:,.0f})")

    if a.flags_csv:
        with open(a.flags_csv, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.DictWriter(fh, fieldnames=["cust", "item", "week", "pkg", "price",
                                               "normal", "dev", "impact", "zero"])
            w.writeheader(); w.writerows(D["red_flags"])
        print(f"דגלים אדומים נשמרו: {a.flags_csv}")
    if a.out:
        json.dump(D, open(a.out, "w", encoding="utf-8"), ensure_ascii=False, separators=(",", ":"))
        print(f"D נכתב: {a.out}")
    else:
        print("\n(יבש — להוספת פלט הוסף --out customers_D.json)")


if __name__ == "__main__":
    main()
