#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
עדכון "מחיר קנייה אחרון" בקטלוג הרכש (לשונית PURCHASING) מכרטסת מלאי שבועית.

מיכל מעלה לתיקייה "מחיר קנייה אחרון" (תחת "קניית ירקות") כרטסת מלאי של השבוע
(פורמט מחשבשבת רב-פריטים). לכל פריט מחלצים את **מחיר הרכש האחרון** — שורת
"חשבונית רכש" האחרונה (מחיר נטו) — ומעדכנים את שדה `p`/`pd` בקטלוג. פריטים
בלי חשבונית רכש בשבוע (רק מכירות/משלוחים) נשארים ללא שינוי.

אבטחה: המחירים נכנסים אך ורק לגיליון, לא לקוד הציבורי. הקטלוג נשמר כ-base64
בעמודה A של לשונית PURCHASING (אותו מנגנון כמו ingest_purchasing.py); שאר
הלשוניות (PO_HISTORY, SUP_MAP, ...) לא נגעות. SUP_MAP מוחל בטעינה בדף — לא כאן.

הרצה:
  python3 update_purch_prices.py <כרטסת_שבועית.xlsx> --key <SA.json>          # יבש
  python3 update_purch_prices.py <כרטסת_שבועית.xlsx> --key <SA.json> --apply  # כתיבה
"""
import argparse
import base64
import io
import json
import re
import sys
import zipfile
from datetime import datetime, date, timedelta

SHEET_ID = "1rWHMhO8zCB8KKzAJwyFYpuKfo_EQ_-rZB8afaiqUv9Q"
TAB = "PURCHASING"
CHUNK = 40000
XL_EPOCH = datetime(1899, 12, 30)          # בסיס תאריכי אקסל


def _load_ws(path):
    """טוען את הלשונית הראשונה. חלק מייצואי מחשבשבת ("מחיר קנייה אחרון")
    מגיעים עם גיליון סגנונות פגום (borderID) והכרזת ממדים חלקית (A1) ששוברים
    את openpyxl ב-read_only. במקרה כזה מנקים את styles.xml (מרובה xf ריקים כדי
    שכל הפניה תיפתר) וקוראים במצב מלא. אחרת קריאה רגילה."""
    from openpyxl import load_workbook
    try:
        return load_workbook(path, read_only=True, data_only=True).active
    except Exception:
        pass
    with zipfile.ZipFile(path, "r") as zin:
        n = 4000
        styles = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                  '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                  '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
                  '<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
                  '<borders count="1"><border/></borders>'
                  '<cellStyleXfs count="1"><xf/></cellStyleXfs>'
                  '<cellXfs count="%d">%s</cellXfs></styleSheet>' % (n, "<xf/>" * n)).encode("utf-8")
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for it in zin.infolist():
                data = zin.read(it.filename)
                if it.filename == "xl/styles.xml":
                    data = styles
                zout.writestr(it, data)
    buf.seek(0)
    return load_workbook(buf, read_only=False, data_only=True).active   # לא read_only: מתעלם מהכרזת A1


def _to_date(v):
    """ערך תא -> datetime (סדרתי-אקסל, datetime, או מחרוזת dd/mm/yy)."""
    if isinstance(v, (datetime, date)):
        return datetime(v.year, v.month, v.day)
    try:
        return XL_EPOCH + timedelta(days=float(v))
    except (TypeError, ValueError):
        return None


def _rows(ws):
    for row in ws.iter_rows(values_only=True):
        yield list(row)


def last_prices_report(path):
    """דוח "מחיר קנייה אחרון חומרי גלם" (עמודות: קוד מיון, מפתח פריט, שם פריט,
    תאריך קניה אחרון, מחיר קניה אחרון לפני הנחות) -> {מפתח: {name,price,date}}.
    מדלגים על מחיר 0 / תאריך 1980 (אין רכש אמיתי) — לא דורסים מחיר קיים בקטלוג."""
    ws = _load_ws(path)
    ki = di = pi = ni = None
    out = {}
    for vals in _rows(ws):
        sv = [str(c).strip() if c is not None else "" for c in vals]
        if ki is None:
            def find(subs):
                for i, s in enumerate(sv):
                    if any(sub in s for sub in subs):
                        return i
                return None
            k_ = find(["מפתח פריט"])
            d_ = find(["תאריך קניה אחרון", "תאריך קנייה אחרון"])
            p_ = find(["מחיר קניה אחרון", "מחיר קנייה אחרון"])
            if k_ is not None and d_ is not None and p_ is not None:
                ki, di, pi, ni = k_, d_, p_, find(["שם פריט"])
            continue
        k = sv[ki] if ki < len(sv) else ""
        if re.fullmatch(r"\d+\.0", k):
            k = k[:-2]
        if not re.fullmatch(r"\d+", k):
            continue
        raw_p = vals[pi] if pi < len(vals) else None
        try:
            price = float(str(raw_p).replace(",", ""))
        except (TypeError, ValueError):
            continue
        dt = _to_date(vals[di] if di < len(vals) else None)
        if price <= 0 or (dt is not None and dt.year <= 1980):
            continue                                  # אין רכש אמיתי — דלג
        name = sv[ni] if (ni is not None and ni < len(sv)) else ""
        out[k] = {"name": name, "price": round(price, 3),
                  "date": dt.strftime("%d/%m/%y") if dt else ""}
    if ki is None:
        sys.exit("לא זוהתה שורת כותרת של דוח 'מחיר קנייה אחרון'.")
    return out


def detect_format(path):
    """'report' = דוח מחיר קנייה אחרון · 'ledger' = כרטסת מלאי (חשבונית רכש)."""
    ws = _load_ws(path)
    for i, vals in enumerate(_rows(ws)):
        joined = " ".join(str(c) for c in vals if c is not None)
        if "מחיר קניה אחרון" in joined or "מחיר קנייה אחרון" in joined:
            return "report"
        if "מחיר נטו" in joined and "כניסה" in joined:
            return "ledger"
        if i > 12:
            break
    return "ledger"


def last_purchase_prices(path):
    """כרטסת מחשבשבת -> {מפתח פריט: {'name','price','date'}} — חשבונית רכש אחרונה."""
    ws = _load_ws(path)
    ci = None
    cur, out = None, {}
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        s0 = str(vals[0] or '').strip()
        s1 = str(vals[1] or '').strip() if len(vals) > 1 else ''
        s2 = str(vals[2] or '').strip() if len(vals) > 2 else ''
        if ci is None:
            sv = [str(c).strip() if c is not None else '' for c in vals]
            if 'מחיר נטו' in sv and 'כניסה' in sv:
                ci = {k: sv.index(k) for k in ['סוג מסמך', 'תאריך', 'מחיר נטו'] if k in sv}
            continue
        if s0.startswith('סה'):
            cur = None
            continue
        if s0 and re.match(r'^\d{4,6}$', s1) and re.match(r'^\d{3}$', s2):
            cur = {'key': s1, 'name': s0}
            continue
        if cur is None:
            continue
        if str(vals[ci['סוג מסמך']] or '').strip() != 'חשבונית רכש':
            continue
        price = vals[ci['מחיר נטו']]
        try:
            price = float(price)
        except (TypeError, ValueError):
            continue
        if price <= 0:
            continue
        d = vals[ci['תאריך']]
        dt = d if isinstance(d, (datetime, date)) else None
        prev = out.get(cur['key'])
        if prev is None or (dt and (prev['dt'] is None or dt >= prev['dt'])):
            out[cur['key']] = {'name': cur['name'], 'price': round(price, 3), 'dt': dt,
                               'date': dt.strftime('%d/%m/%y') if dt else ''}
    return out


def _decode_cat(s):
    """קטלוג נשמר כ-base64; נופלים לקריאת JSON גולמי אם אינו base64."""
    try:
        return json.loads(base64.b64decode(s).decode("utf-8"))
    except Exception:
        return json.loads(s)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("ledger", help="כרטסת מלאי שבועית או דוח 'מחיר קנייה אחרון' (xlsx)")
    ap.add_argument("--key", required=True, help="מפתח חשבון שירות (JSON)")
    ap.add_argument("--sheet", default=SHEET_ID)
    ap.add_argument("--format", choices=["auto", "report", "ledger"], default="auto",
                    help="report=דוח מחיר קנייה אחרון · ledger=כרטסת מלאי · auto=זיהוי אוטומטי")
    ap.add_argument("--apply", action="store_true", help="כתיבה בפועל (אחרת יבש)")
    args = ap.parse_args()

    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    scope = "https://www.googleapis.com/auth/spreadsheets" if args.apply \
        else "https://www.googleapis.com/auth/spreadsheets.readonly"
    creds = Credentials.from_service_account_file(args.key, scopes=[scope])
    sh = build("sheets", "v4", credentials=creds, cache_discovery=False).spreadsheets()

    fmt = detect_format(args.ledger) if args.format == "auto" else args.format
    prices = last_prices_report(args.ledger) if fmt == "report" \
        else last_purchase_prices(args.ledger)
    print(f"פורמט: {'דוח מחיר קנייה אחרון' if fmt == 'report' else 'כרטסת מלאי'}")
    if not prices:
        sys.exit("לא נמצאו מחירי רכש בקובץ.")

    rows = sh.values().get(spreadsheetId=args.sheet, range=f"{TAB}!A:A").execute().get("values", [])
    cat = _decode_cat("".join(r[0] for r in rows if r))
    bykey = {str(it["k"]): it for it in cat["items"]}

    changed, missing = [], []
    for k, v in prices.items():
        it = bykey.get(str(k))
        if not it:
            missing.append((k, v["name"]))
            continue
        old = it.get("p")
        it["p"] = v["price"]
        it["pd"] = v["date"]
        changed.append((v["name"], old, v["price"]))

    print(f"פריטים שנקנו בכרטסת: {len(prices)} | עודכנו בקטלוג: {len(changed)} | לא בקטלוג: {len(missing)}")
    for n, o, p in changed:
        print(f"  {n[:30]:<32} {str(o):>9} → ₪{p}")
    for k, n in missing:
        print(f"  ⚠️ {k} {n} — לא בקטלוג (דלג)")
    if not args.apply:
        print("\n(הרצה יבשה — להוספת --apply לכתיבה)")
        return
    if not changed:
        print("אין מה לעדכן.")
        return

    cat["pricesUpdatedAt"] = date.today().isoformat()
    js = json.dumps(cat, ensure_ascii=False, separators=(",", ":"))
    b64 = base64.b64encode(js.encode("utf-8")).decode("ascii")   # הקטלוג נשמר כ-base64
    chunks = [b64[i:i + CHUNK] for i in range(0, len(b64), CHUNK)]
    sh.values().clear(spreadsheetId=args.sheet, range=f"{TAB}!A:A").execute()
    sh.values().update(spreadsheetId=args.sheet, range=f"{TAB}!A1",
                       valueInputOption="RAW", body={"values": [[c] for c in chunks]}).execute()
    back = sh.values().get(spreadsheetId=args.sheet, range=f"{TAB}!A:A").execute().get("values", [])
    ok = _decode_cat("".join(r[0] for r in back if r)) == cat
    print(f"\nנכתב לקטלוג ({len(chunks)} שורות). אימות הלוך-ושוב: {'תקין' if ok else 'נכשל'}")
    if not ok:
        sys.exit(1)


if __name__ == "__main__":
    main()
